from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QLabel, QVBoxLayout, QPushButton, QHBoxLayout, QLineEdit, QGridLayout, QRadioButton, QButtonGroup, QFrame, QMessageBox
import os
import datetime
from PyQt5.QtGui import QPixmap, QImage
import cv2
import threading
import sys
from concurrent.futures import ThreadPoolExecutor
from PyQt5.QtCore import QTimer, Qt
import openpyxl
class DoctorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        # Initialize your variables
        self.enter_slide_name_entry = None
        self.enter_patch_number_entry = None
        self.enter_medical_record_entry = None
        self.slid_x_record_entry = None
        self.slide_y_record_entry = None
        self.fc2_checkbox = None
        self.image_size_group = None
        self.Lukemiya_type_group = None

        # Initialize UI and cameras
        self.init_ui()
        

        # Create Excel workbook and add data sheet if not exists
        

        # Connect the button click event to the function
        self.capture_image_button.clicked.connect(self.save_data_excl)
        self.init_cameras()
        #self.capture_image_button2.clicked.connect(self.mobile_image_save)
        self.thread_pool = ThreadPoolExecutor(max_workers=4)

        # Initialize cameras and threads
        self.init_cameras()


    # ... (Other methods)

    def create_excel_workbook(self):
        import os
        from openpyxl import Workbook
        slide_name = self.enter_slide_name_entry.text().strip()
        # Create a 'data' directory if not exists
        #if not os.path.exists('data'):
            #os.makedirs('data')
        
        
        self.excel_path = os.path.join("low_"+slide_name,'data.xlsx')
        
        # Create workbook and add data sheet if not exists
        if not os.path.exists(self.excel_path):
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = 'Data'
            header = ['Slide Name', 'Patch Number', 'Medical Record #', 'Flow Cytometry', 'Image Size', 'Leukemia Type','slide_x_axis','slide_y_axis']
            sheet.append(header)
            workbook.save(self.excel_path)
        else:
            workbook = Workbook()
            workbook = openpyxl.load_workbook(self.excel_path)
        
        return workbook

    def save_data_excl(self):
        # Get the values from the input layout
        slide_name = self.enter_slide_name_entry.text().strip()
        patch_number = self.enter_patch_number_entry.text().strip()
        medical_record = self.enter_medical_record_entry.text().strip()
        slide_x_record = self.slide_x_record_entry.text().strip()
        slide_y_record = self.slide_y_record_entry.text().strip()
        flow_cytometry = "Yes" if self.fc2_checkbox.isChecked() else "No"
        
        if not slide_name:
            QMessageBox.critical(self, "Error", "Please enter a slide name.")
            return
        if not patch_number:
            QMessageBox.critical(self, "Error", "Please enter a slide name.")
            return
        if not slide_x_record:
            QMessageBox.critical(self, "Error", "Please enter a slide x_axis.")
            return
        if not slide_y_record:
            QMessageBox.critical(self, "Error", "Please enter a slide y_axis.")
            return
        
        selected_image_size = [checkbox.text() for checkbox in self.image_size_group.buttons() if checkbox.isChecked()]
        if not selected_image_size:
            QMessageBox.critical(self, "Error", "Please select an image size.")
            return
        #selected_image_size=selected_image_size[0]
        leukemia_type = [checkbox.text() for checkbox in self.Lukemiya_type_group.buttons() if checkbox.isChecked()]
        if not leukemia_type:
            QMessageBox.critical(self, "Error", "Please select a leukemia_type.")
            return
        ret, frame = self.camera4.read()
        ret2, frame2 = self.camera.read()
        ret3, frame3 = self.camera2.read()
        
        
        if ret:
            now = datetime.datetime.now()
            date_time_str = now.strftime("%Y-%m-%d_%H-%M-%S")
            folder_name = f"low_{slide_name}"
            slide_res= f"{selected_image_size[0]}"
            image_name= f"Mobile_{slide_name}_{patch_number}_{selected_image_size[0]}_{leukemia_type[0]}"
            
            # Create folder if it doesn't exist
            if not os.path.exists(folder_name):
                os.makedirs(folder_name)
            if not os.path.exists(os.path.join(folder_name,slide_res)):
                os.makedirs(os.path.join(folder_name,slide_res))
            
            
            # Save the image in the folder
            image_path = os.path.join(folder_name,slide_res,image_name+".png")
            image_path2 = os.path.join(folder_name,slide_res,image_name+".jpg")
            image_path3 = os.path.join(folder_name,slide_res,image_name+".jpeg")
            #frame = cv2.rotate(frame, cv2.ROTATE_180)
            cv2.imwrite(image_path, frame)
            cv2.imwrite(image_path2, frame2)
            cv2.imwrite(image_path3, frame3)

            print("Microscpe Image saved!")
            print("Patch number =  " , patch_number)
            print("y_axis coord =  " , slide_y_record)
        # Add the data to the Excel sheet
        self.workbook = self.create_excel_workbook()
        slide_save_name= f"{slide_name}_{patch_number}_{selected_image_size[0]}_{leukemia_type[0]}"
        sheet = self.workbook['Data']
        data_row = [slide_save_name, patch_number, medical_record, flow_cytometry, selected_image_size[0], leukemia_type[0],slide_x_record,slide_y_record]
        
        sheet.append(data_row)
        self.workbook.save(self.excel_path)

        print("Data saved!")
    # capture mobile image and save 
    def mobile_image_save(self):
        # Get the values from the input layout
        slide_name = self.enter_slide_name_entry.text().strip()
        patch_number = self.enter_patch_number_entry.text().strip()
        medical_record = self.enter_medical_record_entry.text().strip()
        flow_cytometry = "Yes" if self.fc2_checkbox.isChecked() else "No" 
        selected_image_size = [checkbox.text() for checkbox in self.image_size_group.buttons() if checkbox.isChecked()]
        #selected_image_size=selected_image_size[0]
        leukemia_type = [checkbox.text() for checkbox in self.Lukemiya_type_group.buttons() if checkbox.isChecked()]
        slide_res= f"{selected_image_size[0]}"
        ret, frame = self.camera4.read()
        if ret:
            now = datetime.datetime.now()
            date_time_str = now.strftime("%Y-%m-%d_%H-%M-%S")
            folder_name = f"{slide_name}"
            image_name= f"Mobile_{slide_name}_{patch_number}_{selected_image_size[0]}_{leukemia_type[0]}"
            # Create folder if it doesn't exist
            if not os.path.exists(folder_name):
                os.makedirs(folder_name)
            if not os.path.exists(os.path.join(folder_name,slide_res)):
                os.makedirs(os.path.join(folder_name,slide_res))

            # Save the image in the folder
            image_path = os.path.join(folder_name,slide_res,image_name+".png")
            cv2.imwrite(image_path, frame)

            print("Mobile Image saved!")
     


    def init_ui(self):
        self.setWindowTitle("Doctor APP")
        self.setGeometry(100, 100, 1500, 1000)

        self.central_widget = QWidget(self)
        self.setCentralWidget(self.central_widget)

        layout = QGridLayout(self.central_widget)

        # Create the input layout
        input_layout = QVBoxLayout()
        input_layout.setSpacing(10)

        
        self.enter_slide_name_label = QLabel("Enter slide Name: ")
        self.enter_slide_name_entry = QLineEdit()
        input_layout.addWidget(self.enter_slide_name_label)
        input_layout.addWidget(self.enter_slide_name_entry)

        self.enter_patch_number_label = QLabel("Enter Patch Number: ")
        self.enter_patch_number_entry = QLineEdit()
        input_layout.addWidget(self.enter_patch_number_label)
        input_layout.addWidget(self.enter_patch_number_entry)

        self.enter_medical_record_label = QLabel("Enter MR#: ")
        self.enter_medical_record_entry = QLineEdit()
        input_layout.addWidget(self.enter_medical_record_label)
        input_layout.addWidget(self.enter_medical_record_entry)

        self.slide_x_record_entry_label = QLabel("Slide x: ")
        self.slide_x_record_entry = QLineEdit()
        input_layout.addWidget(self.slide_x_record_entry_label)
        input_layout.addWidget(self.slide_x_record_entry)

        self.slide_y_record_entry_label = QLabel("Slide y: ")
        self.slide_y_record_entry = QLineEdit()
        input_layout.addWidget(self.slide_y_record_entry_label)
        input_layout.addWidget(self.slide_y_record_entry)
        



        flow_cytometry_group = QButtonGroup(self)
        self.fc1_checkbox = QRadioButton("No")
        self.fc2_checkbox = QRadioButton("Yes")
        flow_cytometry_group.addButton(self.fc1_checkbox)
        flow_cytometry_group.addButton(self.fc2_checkbox)
        flow_cytometry_layout = QHBoxLayout()
        flow_cytometry_layout.addWidget(QLabel("Flow Cytometry:"))
        flow_cytometry_layout.addWidget(self.fc1_checkbox)
        flow_cytometry_layout.addWidget(self.fc2_checkbox)
        input_layout.addLayout(flow_cytometry_layout)
 

        self.image_size_group = QButtonGroup(self)  # Add 'self.' to make it an instance variable
        is1_checkbox = QRadioButton("100")
        is2_checkbox = QRadioButton("400")
        is3_checkbox = QRadioButton("1000")
        self.image_size_group.addButton(is1_checkbox)
        self.image_size_group.addButton(is2_checkbox)
        self.image_size_group.addButton(is3_checkbox)

        image_size_layout = QHBoxLayout()
        image_size_layout.addWidget(QLabel("Image resolution:"))
        image_size_layout.addWidget(is1_checkbox)
        image_size_layout.addWidget(is2_checkbox)
        image_size_layout.addWidget(is3_checkbox)
        input_layout.addLayout(image_size_layout)
        
        self.Lukemiya_type_group = QButtonGroup(self)
        lt1_checkbox = QRadioButton("ALL")
        lt2_checkbox = QRadioButton("AML")
        lt3_checkbox = QRadioButton("CLL")
        lt4_checkbox = QRadioButton("CML")
        lt5_checkbox = QRadioButton("PLL")
        lt6_checkbox = QRadioButton("LGL")
        lt7_checkbox = QRadioButton("HCL")
        lt8_checkbox = QRadioButton("MDS")
        lt9_checkbox = QRadioButton("MYELOMA")
        lt10_checkbox = QRadioButton("NONE")
        lt11_checkbox = QRadioButton("OTHER")
        self.Lukemiya_type_group.addButton(lt1_checkbox)
        self.Lukemiya_type_group.addButton(lt2_checkbox)
        self.Lukemiya_type_group.addButton(lt3_checkbox)
        self.Lukemiya_type_group.addButton(lt4_checkbox)
        self.Lukemiya_type_group.addButton(lt5_checkbox)
        self.Lukemiya_type_group.addButton(lt6_checkbox)
        self.Lukemiya_type_group.addButton(lt7_checkbox)
        self.Lukemiya_type_group.addButton(lt8_checkbox)
        self.Lukemiya_type_group.addButton(lt9_checkbox)
        self.Lukemiya_type_group.addButton(lt10_checkbox)
        self.Lukemiya_type_group.addButton(lt11_checkbox)

        Lukemiya_type_layout = QHBoxLayout()
        Lukemiya_type_layout.addWidget(QLabel("lukemia Type:"))
        Lukemiya_type_layout.addWidget(lt1_checkbox)
        Lukemiya_type_layout.addWidget(lt2_checkbox)
        Lukemiya_type_layout.addWidget(lt3_checkbox)
        Lukemiya_type_layout.addWidget(lt4_checkbox)
        Lukemiya_type_layout.addWidget(lt5_checkbox)
        Lukemiya_type_layout.addWidget(lt6_checkbox)
        Lukemiya_type_layout.addWidget(lt7_checkbox)
        Lukemiya_type_layout.addWidget(lt8_checkbox)
        Lukemiya_type_layout.addWidget(lt9_checkbox)
        Lukemiya_type_layout.addWidget(lt10_checkbox)
        Lukemiya_type_layout.addWidget(lt11_checkbox)
        input_layout.addLayout(Lukemiya_type_layout)

        capture_buttons_frame = QFrame()
        capture_buttons_layout = QHBoxLayout(capture_buttons_frame)

        self.capture_image_button = QPushButton("Mobile  Capture", self)
        capture_buttons_layout.addWidget(self.capture_image_button)

        #self.capture_image_button2 = QPushButton("Mobile Capture", self)
        #capture_buttons_layout.addWidget(self.capture_image_button2)

        input_layout.addWidget(capture_buttons_frame)

        layout.addLayout(input_layout, 0, 0)
        
        
     
        # Add to row 0, column 0
        # Create a horizontal layout for the camera feeds
        # Add to row 0, column 0
        # >>>>>>>>>>>> Add the First column to the main layout
        camera_layout = QHBoxLayout()
        self.live_feed_label2 = QLabel(self)
        camera_layout.addWidget(self.live_feed_label2)

        layout.addLayout(camera_layout, 0, 1)

        #>>>>>>>>>>>> Add the second column to the main layout
        
        camera_layout2 = QHBoxLayout()
        self.live_feed_label4 = QLabel(self)
        camera_layout2.addWidget(self.live_feed_label4)
        layout.addLayout(camera_layout2,1,0)
        camera_layout3 = QHBoxLayout()
        
        #>>>>>>>>>>>> Add the Third column to the main layout
   
        self.live_feed_label = QLabel(self)  # Add this label
        camera_layout3.addWidget(self.live_feed_label)

        #self.live_feed_label3 = QLabel(self)  # Add this label
        #camera_layout3.addWidget(self.live_feed_label3)
        layout.addLayout(camera_layout3,1,1)

     
        
        
       
        

    def init_cameras(self):
        self.camera = cv2.VideoCapture(2)  # Change camera index as needed
        self.camera2 = cv2.VideoCapture(0)
        #self.camera3 = cv2.VideoCapture(2)
        capture_url1 = "http://192.168.42.129:8080/video"
        self.camera4 = cv2.VideoCapture(capture_url1)
        #capture_url2 = "http://192.168.43.1:8080/video"
        #self.camera3 = cv2.VideoCapture(capture_url2)

        self.timer1 = QTimer(self)
        self.timer1.timeout.connect(self.update_live_feed1)
        self.timer1.start(30)  # Set the desired update interval in milliseconds

        self.timer2 = QTimer(self)
        self.timer2.timeout.connect(self.update_live_feed2)
        self.timer2.start(30)

        #self.timer3 = QTimer(self)
        #self.timer3.timeout.connect(self.update_live_feed3)
        #self.timer3.start(30)

        self.timer4 = QTimer(self)
        self.timer4.timeout.connect(self.update_live_feed4)
        self.timer4.start(30)

    def update_live_feed1(self):
        ret, frame = self.camera.read()
        if ret:
            self.update_live_feed(frame, self.live_feed_label)

    def update_live_feed2(self):
        ret, frame = self.camera2.read()
        if ret:
            self.update_live_feed_without_resize(frame, self.live_feed_label2)

    #ef update_live_feed3(self):
        #ret, frame = self.camera3.read()
       # if ret:
            #self.update_live_feed(frame, self.live_feed_label3)

    def update_live_feed4(self):
        ret, frame = self.camera4.read()
        if ret:
            self.update_live_feed(frame, self.live_feed_label4)

    def update_live_feed(self, frame, live_feed_label):
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        frame_rgb = cv2.resize(frame, (360, 300))
        h, w, ch = frame_rgb.shape
        q_image = QImage(frame_rgb.data, w, h, ch * w, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(q_image)
        live_feed_label.setPixmap(pixmap)
   
    def update_live_feed_without_resize (self, frame, live_feed_label):
        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        #frame = cv2.rotate(frame, cv2.ROTATE_180)
        frame_rgb = frame
        h, w, ch = frame_rgb.shape
        q_image = QImage(frame_rgb.data, w, h, ch * w, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(q_image)
        live_feed_label.setPixmap(pixmap)

def main():
    app = QApplication(sys.argv)
    window = DoctorApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
