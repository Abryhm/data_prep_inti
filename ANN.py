import tkinter as tk
from openpyxl import Workbook, load_workbook

from openpyxl.drawing.image import Image as ExcelImage
import cv2
from PIL import Image, ImageTk
import datetime
import os
import time
from tkinter import messagebox
import numpy as np
from PIL import ImageDraw

#import filedialog, messagebox


def show_error_message(title, message):
    messagebox.showerror(title, message)
    

    
def save_data_to_excel(x1,y1,x2,y2,slidename,patch_number,image):   
    
        
        
         
        #Leukemia_types_value = Leukemia_types_var.get()
        #if Leukemia_types_var.get() == "Other":
           # if other_entry.get() == "":
            #    # If checkbox is checked but MR number is not entered, display an error message
             #   show_error_message("Error", "Please enter the other value")
              #  print("Please enter the other value.",flush=True  )
          #  else:
         #       Leukemia_types_value= other_entry.get()
        
        cell_type_value= Cell_types_var.get()        
        if Cell_types_var.get() == "Other":
            if Cell_other_entry.get() == "":
                # If checkbox is checked but MR number is not entered, display an error message
                show_error_message("Error", "Please enter the other cell value")
                print("Please enter the other cell.",flush=True  )
            else:
                cell_type_value= Cell_other_entry.get()
     
                
        
        
        
        #sub5_value = other_entry.get()
        
        attr_size_value = attr_size_var.get()
        attr_Nuclear_Chromation_value = attr_Fairly_homogeneous_var.get()
        attr_Nuclear_shape_value = attr_Nuclear_shape_var.get()
        attr_Nucleous_value = attr_Nucleous_var.get()
        attr_Cytoplasm_value = attr_Cytoplasm_var.get()
        attr_Cytoplasmic_basophilia_value = attr_Cytoplasmic_basophilia_var.get()
        attr_Cytoplasmic_vacuoles_value = attr_Cytoplasmic_vacuoles_var.get()

        
        
        
        
     

        # Open the Excel file
        
        slide_name=slidename.split(os.sep)[1]
        worksheet_name=slidename.split(os.sep)[0]
        
        #>>>>>>>>
        
        
        excel_file_path = os.path.join(worksheet_name,slide_name, slide_name+'_data.xlsx')
        if os.path.exists(excel_file_path):
        # Load the existing workbook
            workbook = load_workbook(excel_file_path)
            sheet = workbook.active
        else:
        # Create a new workbook and add a worksheet
            workbook = Workbook()
            sheet = workbook.active
            #worksheet.title = worksheet_name  # Set the worksheet title
            # Write column headers
            headers = ["Name", "Cell_number", "Leukemia Types", "Cell_type", "Cell Size",
                       "Nuclear Chromation", "Nuclear shape", "Nucleous", "Cytoplasm",
                       "Cytoplasmic basophilia", "Cytoplasmic vacuoles", "X1_cord",
                       "Y1_cord", "X2_cord", "Y2_cord", "Image"]
            sheet.append(headers)
    
        # Select the active worksheet
        #worksheet = workbook[worksheet_name]
        
        #>>>>>>>>>
        #print(worksheet_name)
        #workbook = load_workbook( os.path.join(slide_name,worksheet_name,'data.xlsx'))

        # Select the active sheet
        #sheet = workbook.active

        # Find the next empty row
        next_row = sheet.max_row + 1
        capture_date = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        # Write the data to the Excel sheet
        
        sheet.cell(row=next_row, column=1).value = slidename.split(os.sep)[2]
        sheet.cell(row=next_row, column=2).value = patch_number
        #sheet.cell(row=next_row, column=3).value = Leukemia_types_value
        
        
       

        sheet.cell(row=next_row, column=4).value = cell_type_value
        sheet.cell(row=next_row, column=5).value = attr_size_value
        sheet.cell(row=next_row, column=6).value = attr_Nuclear_Chromation_value
        sheet.cell(row=next_row, column=7).value = attr_Nuclear_shape_value
        sheet.cell(row=next_row, column=8).value = attr_Nucleous_value
        sheet.cell(row=next_row, column=9).value = attr_Cytoplasm_value
        sheet.cell(row=next_row, column=10).value = attr_Cytoplasmic_basophilia_value
        sheet.cell(row=next_row, column=11).value = attr_Cytoplasmic_vacuoles_value
     
        
        #sheet.cell(row=next_row, column=12).value = flow_cytometry_1_value
        #sheet.cell(row=next_row, column=13).value = MR_no_entry_value
        sheet.cell(row=next_row, column=12).value = x1
        sheet.cell(row=next_row, column=13).value = y1
        sheet.cell(row=next_row, column=14).value = x2
        sheet.cell(row=next_row, column=15).value = y2
        img = ExcelImage(image)
        img.anchor = sheet.cell(row=next_row, column=16).coordinate
        sheet.add_image(img)
        
        
        
        
        

        

        # Save the changes to the Excel file
        workbook.save(excel_file_path)
        print("Data saved to data.xlsx",excel_file_path)
        print("Enter button clicked")  
        
    
    
    
    
    
    


    
    
    
    
    
    


# Create the main window
root = tk.Tk()
root.title("Doctor APP")

# Create the main frame
main_frame = tk.Frame(root, bg="lightgray")
main_frame.pack(fill="both", expand=True,anchor="w")

# Create four subframes within the main frame
sub_frame1 = tk.Frame(main_frame, bg="white", width=400, height=480)
sub_frame1.grid(row=0,column=1,rowspan=3, padx=5, pady=5)

sub_frame2 = tk.Frame(main_frame, bg="white", width=800, height=448)
sub_frame2.grid(row=1, column=0, padx=5, pady=5)

sub_frame3 = tk.Frame(main_frame, bg="white", width=600, height=40)
sub_frame3.grid(row=2, column=0, padx=5, pady=5)

sub_frame4 = tk.Frame(main_frame, bg="white", width=600, height=40)
sub_frame4.grid(row=0, column=0, padx=5, pady=5)


def update_frame_border_color1(*args):
    if Cell_types_var.get() != "0":
        Cell_types_frame.config(borderwidth=2, relief="solid", highlightthickness=2, highlightbackground="red")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")
    else:

        Cell_types_frame.config(borderwidth=2, relief="solid", highlightbackground="black")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")


Cell_types_frame = tk.LabelFrame(sub_frame1, text="Cell Types",font=('Times New Roman', 8, 'bold'))
Cell_types_frame.grid(row=3, column=0, columnspan=3, padx=5, pady=5,sticky = 'w')

Cell_types_var = tk.StringVar(value="None")
Cell_types_var .trace_add("write", update_frame_border_color1)

Myeloblast_checkbox = tk.Radiobutton(Cell_types_frame, text="MB", variable=Cell_types_var, value="Myeloblast")
Myeloblast_checkbox.pack(side="left", padx=2, pady=2)


Neutrophil_checkbox = tk.Radiobutton(Cell_types_frame, text="N", variable=Cell_types_var, value="Neutrophil")
Neutrophil_checkbox.pack(side="left", padx=2, pady=2)


basophil_checkbox = tk.Radiobutton(Cell_types_frame, text="B", variable=Cell_types_var, value="Basophil")
basophil_checkbox.pack(side="left", padx=2, pady=2)


Eosinophill_checkbox = tk.Radiobutton(Cell_types_frame, text="E",variable=Cell_types_var, value="Eosinophil")
Eosinophill_checkbox.pack(side="left", padx=2, pady=2)

Myeloma_Plasma_checkbox = tk.Radiobutton(Cell_types_frame, text="M",variable=Cell_types_var, value="Monocyte")
Myeloma_Plasma_checkbox.pack(side="left", padx=2, pady=2)


Lymphoblast_checkbox = tk.Radiobutton(Cell_types_frame, text="LB",variable=Cell_types_var, value="Lymphoblast")
Lymphoblast_checkbox.pack(side="left", padx=5, pady=5)


B_lymphocyte_checkbox = tk.Radiobutton(Cell_types_frame, text="L",variable=Cell_types_var, value="lymphocyte")
B_lymphocyte_checkbox.pack(side="left", padx=2, pady=2)



cell_none_checkbox = tk.Radiobutton(Cell_types_frame, text="None",variable=Cell_types_var, value="None")
cell_none_checkbox.pack(side="left", padx=2, pady=2)

Cell_other = tk.Radiobutton(Cell_types_frame, text="Other:",variable=Cell_types_var, value="Other")
Cell_other.pack(side="left", padx=5, pady=5)

Cell_other_entry = tk.Entry(Cell_types_frame)
Cell_other_entry.pack(side="left", padx=2, pady=2)

 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>



def update_frame_border_color2(*args):
    if attr_size_var.get() != "0":
        Leukemia_Cell_Size_frame.config(borderwidth=2, relief="solid", highlightthickness=2, highlightbackground="red")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")
    else:

        Leukemia_Cell_Size_frame.config(borderwidth=2, relief="solid", highlightbackground="black")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")



Leukemia_Cell_Size_frame = tk.LabelFrame(sub_frame1, text="Cell Size", font=('Times New Roman', 8, 'bold'))
Leukemia_Cell_Size_frame.grid(row=4, column=0, columnspan=3, padx=10, pady=10, sticky='w')

attr_size_var = tk.IntVar(value=0)
attr_size_var.trace_add("write", update_frame_border_color2)

a_radiobutton = tk.Radiobutton(Leukemia_Cell_Size_frame, text="Small", variable=attr_size_var, value=1)
a_radiobutton.pack(side="left", padx=5, pady=5)

b_radiobutton = tk.Radiobutton(Leukemia_Cell_Size_frame, text="Medium", variable=attr_size_var, value=2)
b_radiobutton.pack(side="left", padx=5, pady=5)

c_radiobutton = tk.Radiobutton(Leukemia_Cell_Size_frame, text="Large", variable=attr_size_var, value=3)
c_radiobutton.pack(side="left", padx=5, pady=5)

cn_checkbox = tk.Radiobutton(Leukemia_Cell_Size_frame, text="None", variable=attr_size_var, value=0)
cn_checkbox.pack(side="left", padx=5, pady=5)



 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

def update_frame_border_color3(*args):
    if attr_Fairly_homogeneous_var.get() != "None":
        Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightthickness=2, highlightbackground="red")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")
    else:

        Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="black")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")


Leukemia_Nuclear_Chromation_frame = tk.LabelFrame(sub_frame1, text="Nuclear Chromation",font=('Times New Roman', 8, 'bold'))
Leukemia_Nuclear_Chromation_frame.grid(row=5, column=0,columnspan=3, padx=10, pady=10,sticky = 'w')
attr_Fairly_homogeneous_var = tk.IntVar(value=0)
attr_Fairly_homogeneous_var.trace_add("write", update_frame_border_color3)


d_checkbox = tk.Radiobutton(Leukemia_Nuclear_Chromation_frame, text="Open",variable=attr_Fairly_homogeneous_var, value=1)
d_checkbox.pack(side="left", padx=5, pady=5)


e_checkbox = tk.Radiobutton(Leukemia_Nuclear_Chromation_frame, text="Coarse",variable=attr_Fairly_homogeneous_var, value=2)
e_checkbox.pack(side="left", padx=5, pady=5)


f_checkbox = tk.Radiobutton(Leukemia_Nuclear_Chromation_frame, text="stippled",variable=attr_Fairly_homogeneous_var, value=3)
f_checkbox.pack(side="left", padx=5, pady=5)

fn_checkbox = tk.Radiobutton(Leukemia_Nuclear_Chromation_frame, text="None",variable=attr_Fairly_homogeneous_var, value=0)
fn_checkbox.pack(side="left", padx=5, pady=5)

 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
def update_frame_border_color4(*args):
    if attr_Nuclear_shape_var.get() != "None":
        Leukemia_Nuclear_shape_frame.config(borderwidth=2, relief="solid", highlightthickness=2, highlightbackground="red")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")
    else:

        Leukemia_Nuclear_shape_frame.config(borderwidth=2, relief="solid", highlightbackground="black")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")


Leukemia_Nuclear_shape_frame = tk.LabelFrame(sub_frame1, text="Nuclear shape",font=('Times New Roman', 8, 'bold'))
Leukemia_Nuclear_shape_frame.grid(row=6, column=0,columnspan=3, padx=10, pady=10,sticky = 'w')
attr_Nuclear_shape_var = tk.IntVar(value= 0)
attr_Nuclear_shape_var.trace_add("write", update_frame_border_color4)
g_checkbox = tk.Radiobutton(Leukemia_Nuclear_shape_frame, text="Regular",variable=attr_Nuclear_shape_var, value=1)
g_checkbox.pack(side="left", padx=5, pady=5)


h_checkbox = tk.Radiobutton(Leukemia_Nuclear_shape_frame, text="Iregular",variable=attr_Nuclear_shape_var,value=2)
h_checkbox.pack(side="left", padx=5, pady=5)


#i_checkbox = tk.Radiobutton(Leukemia_Nuclear_shape_frame, text="Regular Over-round",variable=attr_Nuclear_shape_var,value=3)
#i_checkbox.pack(side="left", padx=5, pady=5)

in_checkbox = tk.Radiobutton(Leukemia_Nuclear_shape_frame, text="None",variable=attr_Nuclear_shape_var,value=0)
in_checkbox.pack(side="left", padx=5, pady=5)
  
 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
def update_frame_border_color5(*args):
    if attr_Nucleous_var.get() != "None":
        Leukemia_Nucleous_frame .config(borderwidth=2, relief="solid", highlightthickness=2, highlightbackground="red")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")
    else:

        Leukemia_Nucleous_frame .config(borderwidth=2, relief="solid", highlightbackground="black")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")
    
Leukemia_Nucleous_frame = tk.LabelFrame(sub_frame1, text="Nucleous",font=('Times New Roman', 8, 'bold'))
Leukemia_Nucleous_frame.grid(row=7, column=0,columnspan=3, padx=10, pady=10,sticky = 'w')
attr_Nucleous_var = tk.IntVar(value=0)
attr_Nucleous_var.trace_add("write", update_frame_border_color5)
j_checkbox = tk.Radiobutton(Leukemia_Nucleous_frame, text="Inconspicuous",variable=attr_Nucleous_var,value=1)
j_checkbox.pack(side="left", padx=5, pady=5)


#k_checkbox = tk.Radiobutton(Leukemia_Nucleous_frame, text="Usually visible, Often large ",variable=attr_Nucleous_var,value=2)
#k_checkbox.pack(side="left", padx=5, pady=5)


l_checkbox = tk.Radiobutton(Leukemia_Nucleous_frame, text="Prominent",variable=attr_Nucleous_var, value=3)
l_checkbox.pack(side="left", padx=5, pady=5)

ln_checkbox = tk.Radiobutton(Leukemia_Nucleous_frame, text="None",variable=attr_Nucleous_var, value=0 )
ln_checkbox.pack(side="left", padx=5, pady=5)

 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
def update_frame_border_color6(*args):
    if attr_Cytoplasm_var.get() != "None":
        Leukemia_Cytoplasm_frame.config(borderwidth=2, relief="solid", highlightthickness=2, highlightbackground="red")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")
    else:

        Leukemia_Cytoplasm_frame.config(borderwidth=2, relief="solid", highlightbackground="black")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")




Leukemia_Cytoplasm_frame = tk.LabelFrame(sub_frame1, text="Cytoplasm",font=('Times New Roman', 8, 'bold'))
Leukemia_Cytoplasm_frame.grid(row=8, column=0,columnspan=3, padx=10, pady=10,sticky = 'w')

attr_Cytoplasm_var = tk.IntVar(value= 0)
attr_Cytoplasm_var.trace_add("write", update_frame_border_color6)
m_checkbox = tk.Radiobutton(Leukemia_Cytoplasm_frame, text="Scanty",variable=attr_Cytoplasm_var, value=1)
m_checkbox.pack(side="left", padx=5, pady=5)



n_checkbox = tk.Radiobutton(Leukemia_Cytoplasm_frame, text="Abundent",variable=attr_Cytoplasm_var, value=2)
n_checkbox.pack(side="left", padx=5, pady=5)


#o_checkbox = tk.Radiobutton(Leukemia_Cytoplasm_frame, text="Moderately abundant",variable=attr_Cytoplasm_var, value =3)
#o_checkbox.pack(side="left", padx=5, pady=5)

on_checkbox = tk.Radiobutton(Leukemia_Cytoplasm_frame, text="None",variable=attr_Cytoplasm_var, value = 0)
on_checkbox.pack(side="left", padx=5, pady=5)


 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
def update_frame_border_color7(*args):
    if attr_Cytoplasmic_basophilia_var.get() != "None":
        Leukemia_Cytoplasmic_basophilia_frame.config(borderwidth=2, relief="solid", highlightthickness=2, highlightbackground="red")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")
    else:

        Leukemia_Cytoplasmic_basophilia_frame.config(borderwidth=2, relief="solid", highlightbackground="black")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")


Leukemia_Cytoplasmic_basophilia_frame = tk.LabelFrame(sub_frame1, text="Cytoplasmic basophilia",font=('Times New Roman', 12, 'bold'))
Leukemia_Cytoplasmic_basophilia_frame.grid(row=9, column=0,columnspan=3, padx=10, pady=10,sticky = 'w')
attr_Cytoplasmic_basophilia_var = tk.IntVar(value=0)
attr_Cytoplasmic_basophilia_var.trace_add("write", update_frame_border_color7)
p_checkbox = tk.Radiobutton(Leukemia_Cytoplasmic_basophilia_frame, text="Slight ",variable=attr_Cytoplasmic_basophilia_var,value =1)
p_checkbox.pack(side="left", padx=5, pady=5)


q_checkbox = tk.Radiobutton(Leukemia_Cytoplasmic_basophilia_frame, text="moderate",variable=attr_Cytoplasmic_basophilia_var, value=2)
q_checkbox.pack(side="left", padx=5, pady=5)


r_checkbox = tk.Radiobutton(Leukemia_Cytoplasmic_basophilia_frame, text="Strong",variable=attr_Cytoplasmic_basophilia_var, value=3)
r_checkbox.pack(side="left", padx=5, pady=5)

rn_checkbox = tk.Radiobutton(Leukemia_Cytoplasmic_basophilia_frame, text="None",variable=attr_Cytoplasmic_basophilia_var, value=0)
rn_checkbox.pack(side="left", padx=5, pady=5)

 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
def update_frame_border_color9(*args):
    if attr_Cytoplasmic_vacuoles_var.get() != "None":
        Leukemia_Cytoplasmic_vacuoles_frame.config(borderwidth=2, relief="solid", highlightthickness=2, highlightbackground="red")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")
    else:

        Leukemia_Cytoplasmic_vacuoles_frame.config(borderwidth=2, relief="solid", highlightbackground="black")
        #Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="red")

Leukemia_Cytoplasmic_vacuoles_frame = tk.LabelFrame(sub_frame1, text="Cytoplasmic vacuoles",font=('Times New Roman', 8, 'bold'))
Leukemia_Cytoplasmic_vacuoles_frame.grid(row=10, column=0,columnspan=2, padx=10, pady=10,sticky = 'w')
attr_Cytoplasmic_vacuoles_var = tk.IntVar(value = 0)
attr_Cytoplasmic_vacuoles_var.trace_add("write", update_frame_border_color9)
s_checkbox = tk.Radiobutton(Leukemia_Cytoplasmic_vacuoles_frame, text="Absent",variable=attr_Cytoplasmic_vacuoles_var,value =1)
s_checkbox.pack(side="left", padx=5, pady=5)


#t_checkbox = tk.Radiobutton(Leukemia_Cytoplasmic_vacuoles_frame, text="Absent",variable=attr_Cytoplasmic_vacuoles_var, value=2)
#t_checkbox.pack(side="left", padx=5, pady=5)


u_checkbox =tk.Radiobutton(Leukemia_Cytoplasmic_vacuoles_frame, text="Prominent",variable=attr_Cytoplasmic_vacuoles_var, value =3)
u_checkbox.pack(side="left", padx=5, pady=5)

u_checkbox =tk.Radiobutton(Leukemia_Cytoplasmic_vacuoles_frame, text="None",variable=attr_Cytoplasmic_vacuoles_var, value =0)
u_checkbox.pack(side="left", padx=5, pady=5)


 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>



def reset_frame_border_color():
    Leukemia_Cell_Size_frame.config(borderwidth=2, relief="solid", highlightbackground="black")
    #attr_size_var.set(0)
    Cell_types_frame.config(borderwidth=2, relief="solid", highlightbackground="black")
    #Cell_types_var.set("None")  
    Leukemia_Nuclear_Chromation_frame.config(borderwidth=2, relief="solid", highlightbackground="black") 
    Leukemia_Cytoplasmic_vacuoles_frame.config(borderwidth=2, relief="solid", highlightbackground="black")
    Leukemia_Cytoplasmic_basophilia_frame.config(borderwidth=2, relief="solid", highlightbackground="black")
    Leukemia_Cytoplasm_frame.config(borderwidth=2, relief="solid", highlightbackground="black")
    Leukemia_Nucleous_frame.config(borderwidth=2, relief="solid", highlightbackground="black")
    Leukemia_Nuclear_shape_frame.config(borderwidth=2, relief="solid", highlightbackground="black")
    #Leukemia_Cytoplasmic_basophilia_frame.config(borderwidth=2, relief="solid", highlightbackground="black")



















 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
image_name_label = tk.Label(sub_frame2, text="", font=('Times New Roman', 15, 'bold'))
image_name_label.pack()
from tkinter import filedialog


idx=0
image_paths = []
current_image_index = -1


bbox_start_x = bbox_start_y = 0
current_drawing_box = None
bounding_boxes = []
previous_drawing_boxes = []
def show_completion_message():
    messagebox.showinfo("Completion", "Folder annotation complete!")
def select_folder():
    global folder_path,filename
    folder_path = filedialog.askdirectory()
    if folder_path:
        global image_paths, current_image_index 
        image_paths = [os.path.join(folder_path, filename) for filename in os.listdir(folder_path) if filename.lower().endswith(('.jpg', '.jpeg', '.png', '.bmp'))]
        
        if not image_paths:
            show_error_message("Error", "No image files found in the selected folder.")
            return
        current_image_index = 0
        show_next_image()
        
def show_next_image():
    global current_image_index, selected_image,bbox_start_x,bbox_start_y,idx, selected_photo,current_drawing_box, bounding_boxes, drawing_box,previous_drawing_boxes
    if current_image_index < len(image_paths):
        slidename = os.path.basename(image_paths[current_image_index])
        image_name_label.config(text=f"Selected Image: {slidename}")

        selected_image = Image.open(image_paths[current_image_index])
        selected_image.thumbnail((800, 448))
        selected_photo = ImageTk.PhotoImage(selected_image)

        # Create an image item on the canvas
        image_item = selected_image_label.create_image(0, 0, anchor="nw", image=selected_photo)
        selected_image_label.image_item = image_item  # Save a reference to prevent garbage collection

        bounding_boxes = []
        previous_drawing_boxes=[]
        drawing_box = None

        current_image_index += 1
        bbox_start_x = bbox_start_y = 0
        current_drawing_box = None
        bounding_boxes = []
        previous_drawing_boxes = []
        idx=0

def start_drawing(event):
    global bbox_start_x, bbox_start_y, current_drawing_box
    bbox_start_x = event.x
    bbox_start_y = event.y
    current_drawing_box = selected_image_label.create_rectangle(
        bbox_start_x, bbox_start_y, bbox_start_x, bbox_start_y, outline="red", width=2
    )

def draw_bounding_box(event):
    global current_drawing_box
    if current_drawing_box:
        x, y = event.x, event.y
        selected_image_label.coords(current_drawing_box, bbox_start_x, bbox_start_y, x, y)

bounding_box_labels = []
def end_drawing(event):
    global current_drawing_box, bounding_boxes, idx, previous_drawing_boxes,bounding_box_labels
    if current_drawing_box:
        x, y = event.x, event.y
        selected_image_label.coords(current_drawing_box, bbox_start_x, bbox_start_y, x, y)
        bounding_boxes=bbox_start_x, bbox_start_y, x, y
        #idx += 1
        label_x = (bbox_start_x - 5) #/ 2
        label_y = (bbox_start_y - 5) #/ 2
        bbox_label = selected_image_label.create_text(label_x, label_y, text=str(idx), fill="red")
        bounding_box_labels.append(bbox_label)
        previous_drawing_boxes.append(current_drawing_box)
        current_drawing_box = None
def save_bounded_areas():
    global bounding_boxes, current_image_index, image_paths,idx
    
    Annotation = "Annotation"
    os.makedirs(Annotation, exist_ok=True)
    folderimage_path= folder_path.split("/")[-1]
    print(folderimage_path)
    folderimage_path= os.path.join(Annotation,folder_path.split("/")[-1])
    os.makedirs(folderimage_path, exist_ok=True)
    
    I_path = image_paths[current_image_index - 1]
    save_image_path = os.path.join(folderimage_path, os.path.splitext(os.path.basename(I_path))[0])
    
    os.makedirs(save_image_path, exist_ok=True)
    
    #for idx, bbox in enumerate(bounding_boxes):
    x1, y1, x2, y2 = bounding_boxes
    bounded_area = selected_image.crop((x1, y1, x2, y2))
    idx +=1
    patch_num = str(idx)
    save_filename = os.path.join(save_image_path, f"{patch_num}.png")
    bounded_area.save(save_filename)
    save_data_to_excel(x1, y1, x2, y2, save_filename, patch_num, save_filename)
        
    # Clear the previous drawing boxes and reset outline color
    for box_id in previous_drawing_boxes:
        selected_image_label.itemconfig(box_id, outline="yellow")
    previous_drawing_boxes.clear()
    reset_frame_border_color()
    




# Create frames and widgets as needed

def clear_last_bounding_box():
    global previous_drawing_boxes,bounding_box_labels
    if previous_drawing_boxes:
        box_to_delete = previous_drawing_boxes.pop()
        selected_image_label.delete(box_to_delete)
        label_to_delete=bounding_box_labels.pop()
        selected_image_label.delete(label_to_delete)
        

      
        




selected_image_label = tk.Canvas(sub_frame2)
selected_image_label.configure(width=800, height=448)
selected_image_label.pack()


select_image_button = tk.Button(sub_frame4, text="Select Folder", command=select_folder, font=('Times New Roman', 15, 'bold'))
select_image_button.pack(side="left", padx=10, pady=10)


#selected_slide_label = tk.Label(sub_frame4, text="Selected Slide: ", font=('Times New Roman', 15, 'bold'))
#selected_slide_label.pack(side="left", padx=10, pady=10)


next_image_button = tk.Button(sub_frame3, text="Next", command=show_next_image, font=('Times New Roman', 15, 'bold'))
next_image_button.pack(side="left", padx=10, pady=10)

# = tk.Button(root, text="Select Folder", command=select_folder)
##select_folder_button.pack()



selected_image_label.bind("<Button-1>", start_drawing)
selected_image_label.bind("<B1-Motion>", draw_bounding_box)
selected_image_label.bind("<ButtonRelease-1>", end_drawing)

save_areas_button = tk.Button(sub_frame1, text="Save", command=save_bounded_areas,font=('Times New Roman', 15, 'bold'))
save_areas_button.grid(row=11, column=1, padx=5, pady=5,sticky = 'w')


clear_last_box_button = tk.Button(sub_frame1, text="Clear Last Bounding Box", command=clear_last_bounding_box, font=('Times New Roman', 15, 'bold'))
clear_last_box_button.grid(row=12, column=1, padx=5, pady=5, sticky='w')









 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
 #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>



root.mainloop()




